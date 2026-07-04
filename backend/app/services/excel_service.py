"""Excel export service."""

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.property import Property


class ExcelService:
    """Service for exporting properties to Excel."""

    async def generate(self, properties: list[Property]) -> io.BytesIO:
        """Generate Excel file from properties."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Недвижимость"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alignment = Alignment(wrap_text=True, vertical="top")

        # Define columns
        columns = [
            ("ID", 10),
            ("Внешний ID", 15),
            ("Название", 40),
            ("Тип", 15),
            ("Тип продажи", 15),
            ("Регион", 20),
            ("Город", 20),
            ("Адрес", 30),
            ("Площадь", 12),
            ("Комнат", 10),
            ("Цена", 15),
            ("Валюта", 10),
            ("Статус", 15),
            ("Дата парсинга", 18),
            ("URL", 30),
        ]

        # Write header
        for col_num, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Write data
        for row_num, prop in enumerate(properties, 2):
            data = [
                prop.id,
                prop.external_id,
                prop.title,
                prop.property_type,
                prop.sale_type,
                prop.region or "",
                prop.city or "",
                prop.address or "",
                prop.total_area,
                prop.rooms,
                float(prop.price) if prop.price else None,
                prop.currency,
                prop.status,
                prop.parsed_at.strftime("%Y-%m-%d %H:%M") if prop.parsed_at else "",
                prop.source_url or "",
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = alignment

                # Format price column
                if col_num == 11 and value:
                    cell.number_format = "#,##0.00"

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary sheet
        ws_summary = wb.create_sheet("Сводка")
        self._add_summary(ws_summary, properties)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _add_summary(self, ws: Any, properties: list[Property]) -> None:
        """Add summary statistics to worksheet."""
        # Styles
        header_font = Font(bold=True, size=14)
        label_font = Font(bold=True)

        # Title
        ws.cell(row=1, column=1, value="Сводная статистика")
        ws.cell(row=1, column=1).font = header_font

        # Date
        ws.cell(row=2, column=1, value=f"Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # Total count
        ws.cell(row=4, column=1, value="Всего объектов:")
        ws.cell(row=4, column=1).font = label_font
        ws.cell(row=4, column=2, value=len(properties))

        # By property type
        ws.cell(row=6, column=1, value="По типу недвижимости:")
        ws.cell(row=6, column=1).font = label_font

        type_counts: dict[str, int] = {}
        for prop in properties:
            type_counts[prop.property_type] = type_counts.get(prop.property_type, 0) + 1

        row = 7
        for prop_type, count in sorted(type_counts.items()):
            ws.cell(row=row, column=1, value=prop_type)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # By sale type
        ws.cell(row=row + 1, column=1, value="По типу продажи:")
        ws.cell(row=row + 1, column=1).font = label_font

        sale_counts: dict[str, int] = {}
        for prop in properties:
            sale_counts[prop.sale_type] = sale_counts.get(prop.sale_type, 0) + 1

        row += 2
        for sale_type, count in sorted(sale_counts.items()):
            ws.cell(row=row, column=1, value=sale_type)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Price statistics
        prices = [float(p.price) for p in properties if p.price]
        if prices:
            row += 1
            ws.cell(row=row, column=1, value="Ценовая статистика:")
            ws.cell(row=row, column=1).font = label_font

            row += 1
            ws.cell(row=row, column=1, value="Минимальная цена:")
            ws.cell(row=row, column=2, value=min(prices))
            ws.cell(row=row, column=2).number_format = "#,##0.00"

            row += 1
            ws.cell(row=row, column=1, value="Максимальная цена:")
            ws.cell(row=row, column=2, value=max(prices))
            ws.cell(row=row, column=2).number_format = "#,##0.00"

            row += 1
            ws.cell(row=row, column=1, value="Средняя цена:")
            ws.cell(row=row, column=2, value=sum(prices) / len(prices))
            ws.cell(row=row, column=2).number_format = "#,##0.00"

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
